#!/usr/bin/env python3
"""
Volume III, step 1: build the question sample and the study universe.

Inputs
  - GEO_Prospect_List_54.xlsx  (Sweep tab: company, domain, category, competitors,
    brand_aliases; Sheet2 tab: the buyer-question bank as `prompt`)
  - state-of-geo-2026 repo     (Volume I / II open dataset)

Never reads any summary column from the xlsx. Questions, categories, company
names, competitors and aliases only. The domain_cited column in that sheet is
Excel-date-corrupted and is not touched.

Outputs
  data/questions.csv        one row per sampled question
  data/universe.csv         one row per company in the matching universe
  data/sample_manifest.json counts and the selection rule
"""
import csv, json, collections, re, os, sys
import openpyxl

XLSX = os.environ.get("VOL3_XLSX",
    "/root/.claude/uploads/2a6b0b0d-bfde-53d3-a976-6afaf2b5f362/b72fbd11-GEO_Prospect_List_54.xlsx")
REPO = "data/state-of-geo-2026"
OUT = "data"

N_CATEGORIES = 40
Q_PER_CATEGORY = 7
STABILITY_N = 25
SHAPES = ["best-of", "alternatives", "comparison", "use-case", "evaluation"]

TEST_NAMES = {"jane doe", "leader test"}
TEST_COMPANIES = {"exampleco", "example co"}


def norm(s):
    return re.sub(r"\s+", " ", str(s or "")).strip()


# ---------------------------------------------------------------- load xlsx
wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb["Sweep"]
hdr = [c.value for c in ws[1]]
sweep_all = [dict(zip(hdr, r)) for r in ws.iter_rows(min_row=2, values_only=True) if r[1]]

# ExampleCo / Jane Doe are dummy rows and are dropped everywhere.
# LEADER TEST rows are category leaders, not Volume I measured subjects. They stay
# in the frame as category carriers and as nameable vendors, but they are never
# measured subjects and never appear in any ranking.
sweep = []
excluded = []
leader_rows = []
for r in sweep_all:
    nm = norm(r.get("name")).lower()
    co = norm(r.get("company")).lower()
    if nm in TEST_COMPANIES or co in TEST_COMPANIES or nm == "jane doe":
        excluded.append((norm(r.get("name")), norm(r.get("company")), norm(r.get("category"))))
        continue
    r["_is_leader_row"] = (nm == "leader test")
    if r["_is_leader_row"]:
        leader_rows.append(norm(r.get("company")))
    sweep.append(r)

w2 = wb["Sheet2"]
h2 = [c.value for c in w2[1]]
raw = [dict(zip(h2, r)) for r in w2.iter_rows(min_row=2, values_only=True) if r[1]]

company_to_cat = {}
for r in sweep:
    c = norm(r.get("company"))
    if c and c not in company_to_cat:
        company_to_cat[c] = norm(r.get("category"))

# ---------------------------------------------------------------- Vol I frame
vol1_cats = sorted({r["category"] for r in
                    csv.DictReader(open(f"{REPO}/challenger_visibility_v2.csv"))})
challengers = list(csv.DictReader(open(f"{REPO}/challenger_visibility_v2.csv")))
companies_per_cat = collections.Counter(r["category"] for r in challengers)

# Vol I labelled question shapes, reused verbatim where the question matches.
vol1_shape = {}
vol1_qs = collections.defaultdict(set)
for r in csv.DictReader(open(f"{REPO}/absence_questions_classified.csv")):
    q = norm(r["question_text"])
    vol1_shape[q] = r["question_shape"]
    vol1_qs[r["category"]].add(q)


def classify(q):
    """Volume I question-shape rules, applied only to questions Volume I did not label."""
    t = q.lower()
    if re.search(r"\bvs\.?\b|\bversus\b|compare[sd]?\b|how do .* compare", t):
        return "comparison"
    if re.search(r"\balternativ|\breplace(ment)?\b|\bcompetitors? to\b|instead of\b", t):
        return "alternatives"
    if re.search(r"\bpric|\bcost\b|\bbudget\b|\broi\b|\bworth it\b|how much\b|"
                 r"\bevaluat|\bchecklist\b|\bcriteria\b|what should i look for|"
                 r"questions to ask\b", t):
        return "evaluation"
    if re.search(r"^what is the best|^which .* (is|are) best|\bbest\b", t):
        return "best-of"
    if re.search(r"\bfor (a|an|my|our|small|mid|large|enterprise|teams?|companies|"
                 r"agencies|startups?)\b|\buse case\b|\bwhen (you|i|we)\b|\bhow (do|can) "
                 r"(i|we|you)\b", t):
        return "use-case"
    return "other"


# ---------------------------------------------------------------- bank per category
bank = collections.defaultdict(dict)   # category -> {question: source}
for r in raw:
    cat = company_to_cat.get(norm(r.get("company")))
    q = norm(r.get("prompt"))
    if cat and q and q not in bank[cat]:
        bank[cat][q] = "question_bank"
for cat, qs in vol1_qs.items():
    for q in qs:
        if q not in bank[cat]:
            bank[cat][q] = "vol1_dataset"

# ---------------------------------------------------------------- pick categories
# Deterministic rule, stated in the paper: every Volume I category carrying more
# than one measured company is included, then singleton categories in alphabetical
# order until N_CATEGORIES is reached. Maximises company-level tie-back coverage.
multi = sorted([c for c in vol1_cats if companies_per_cat[c] > 1],
               key=lambda c: (-companies_per_cat[c], c))
singles = sorted([c for c in vol1_cats if companies_per_cat[c] == 1])
chosen = (multi + singles)[:N_CATEGORIES]
chosen = [c for c in chosen if len(bank[c]) >= Q_PER_CATEGORY]
assert len(chosen) == N_CATEGORIES, f"only {len(chosen)} usable categories"

# ---------------------------------------------------------------- pick questions
questions = []
qid = 0
shape_short = []
for cat in sorted(chosen):
    pool = sorted(bank[cat].keys())
    shaped = collections.defaultdict(list)
    for q in pool:
        s = vol1_shape.get(q) or classify(q)
        shaped[s].append(q)
    picked, seen = [], set()
    # one of each required shape first
    for s in SHAPES:
        for q in shaped.get(s, []):
            if q not in seen:
                picked.append((q, s)); seen.add(q); break
    missing = [s for s in SHAPES if s not in {p[1] for p in picked}]
    if missing:
        shape_short.append({"category": cat, "missing_shapes": missing})
    # fill to Q_PER_CATEGORY round-robin across shapes so no single shape dominates
    queues = {s: [q for q in shaped.get(s, []) if q not in seen] for s in SHAPES}
    queues["other"] = [q for q in shaped.get("other", []) if q not in seen]
    ring = SHAPES + ["other"]
    i = 0
    while len(picked) < Q_PER_CATEGORY and any(queues[s] for s in ring):
        s = ring[i % len(ring)]
        i += 1
        if queues[s]:
            q = queues[s].pop(0)
            if q not in seen:
                picked.append((q, s)); seen.add(q)
    for q, s in picked[:Q_PER_CATEGORY]:
        qid += 1
        questions.append({
            "question_id": f"Q{qid:04d}",
            "category": cat,
            "question_text": q,
            "question_shape": s,
            "source": bank[cat][q],
            "shape_source": "vol1_labelled" if q in vol1_shape else "regenerated_label",
        })

# ---------------------------------------------------------------- stability subsample
# 25 questions, stratified: evenly spaced across the sorted category list, one per
# category until 25 are taken, shape-balanced by taking each category's i-th shape.
cats_sorted = sorted(chosen)
stability = []
step = len(cats_sorted) / STABILITY_N
by_cat = collections.defaultdict(list)
for q in questions:
    by_cat[q["category"]].append(q)
for i in range(STABILITY_N):
    cat = cats_sorted[int(i * step)]
    pool = [q for q in by_cat[cat] if q["question_id"] not in {s["question_id"] for s in stability}]
    pool.sort(key=lambda q: (SHAPES.index(q["question_shape"]) if q["question_shape"] in SHAPES else 99,
                             q["question_id"]))
    stability.append(pool[i % len(pool)])
stab_ids = {q["question_id"] for q in stability}
for q in questions:
    q["in_stability_subsample"] = q["question_id"] in stab_ids

# ---------------------------------------------------------------- universe
universe = []
for r in sweep:
    co = norm(r.get("company"))
    if not co:
        continue
    universe.append({
        "company": co,
        "domain": norm(r.get("domain")),
        "category": norm(r.get("category")),
        "competitors": norm(r.get("competitors")),
        "brand_aliases": norm(r.get("brand_aliases")),
        "in_sampled_category": norm(r.get("category")) in set(chosen),
        "measured_subject": not r.get("_is_leader_row", False),
    })
# competitor names are part of the matching universe too
comp_names = set()
for u in universe:
    for c in re.split(r"\s*,\s*", u["competitors"]):
        c = norm(c)
        if c:
            comp_names.add(c)

os.makedirs(OUT, exist_ok=True)
with open(f"{OUT}/questions.csv", "w", newline="") as f:
    w = csv.DictWriter(f, fieldnames=list(questions[0].keys()))
    w.writeheader(); w.writerows(questions)
with open(f"{OUT}/universe.csv", "w", newline="") as f:
    w = csv.DictWriter(f, fieldnames=list(universe[0].keys()))
    w.writeheader(); w.writerows(universe)

manifest = {
    "n_categories": len(chosen),
    "questions_per_category": Q_PER_CATEGORY,
    "n_questions": len(questions),
    "n_stability_questions": len(stability),
    "stability_repeats": 3,
    "engines": ["chatgpt", "claude", "perplexity", "google_aio"],
    "planned_main_calls": len(questions) * 4,
    "planned_stability_extra_calls": len(stability) * 2 * 4,
    "serpapi_worst_case": (len(questions) + len(stability) * 2) * 2,
    "category_selection_rule": ("all Volume I categories carrying more than one measured "
                                "company, then singleton categories alphabetically, to 40"),
    "excluded_test_rows": excluded,
    "leader_rows_not_measured": sorted(set(leader_rows)),
    "categories": sorted(chosen),
    "shape_coverage_shortfalls": shape_short,
    "universe_companies": len(universe),
    "universe_competitor_names": len(comp_names),
    "question_shape_mix": dict(collections.Counter(q["question_shape"] for q in questions)),
    "question_source_mix": dict(collections.Counter(q["source"] for q in questions)),
}
json.dump(manifest, open(f"{OUT}/sample_manifest.json", "w"), indent=2)
print(json.dumps({k: v for k, v in manifest.items()
                  if k not in ("categories", "excluded_test_rows")}, indent=2))
print("shape shortfalls:", shape_short)
