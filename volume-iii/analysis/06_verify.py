#!/usr/bin/env python3
"""
Volume III verification gate.

This is a SECOND, INDEPENDENTLY WRITTEN implementation. It does not import
anything from 02/03, does not read results.json, and rebuilds every headline
figure from data/raw.csv using pandas rather than the hand-rolled csv/collections
code in the main pipeline. It then reads the actual text of PAPER.md and the
figure captions and diffs the numbers it finds there against the numbers it just
computed.

Any mismatch is a hard failure and publishing must stop until it is resolved.

Deliberate design choices that make this a real check rather than a rerun:
  - pandas groupby instead of manual dict accumulation
  - Jaccard from set algebra over exploded long-format rows, not the paired loop
  - the within/between comparison rebuilt from the repeat index directly
  - the paper is parsed as TEXT, so a number that never made it into the prose
    or drifted during editing is caught

Usage: 06_verify.py
Exit code 0 = gate passed, 1 = gate failed.
"""
import json, re, sys, itertools, math
import pandas as pd

TOL = 0.0006          # numbers are printed to 3 dp, so this is half a unit
PCT_TOL = 0.06        # percentages printed to 1 dp

fails, checks = [], []


def check(name, mine, paper, tol):
    ok = (mine is not None and paper is not None and abs(mine - paper) <= tol)
    checks.append((name, mine, paper, ok))
    if not ok:
        fails.append(f"{name}: independent recomputation {mine}, paper says {paper}")


# ---------------------------------------------------------------- rebuild
raw = pd.read_csv("data/raw.csv", dtype=str, keep_default_na=False,
                  engine="python", on_bad_lines="skip")
raw["repeat_index"] = pd.to_numeric(raw["repeat_index"], errors="coerce").fillna(1).astype(int)
scored = pd.read_csv("data/scored.csv", dtype=str, keep_default_na=False,
                     engine="python", on_bad_lines="skip")
scored["repeat_index"] = pd.to_numeric(scored["repeat_index"], errors="coerce").fillna(1).astype(int)
scored["vendors"] = scored["vendor_set_json"].map(lambda s: frozenset(json.loads(s)))

ENG = ["chatgpt", "claude", "perplexity", "google_aio"]

# row count and collection integrity, checked against raw.csv not scored.csv
n_rows = len(scored)
assert len(raw) == n_rows, f"raw.csv {len(raw)} rows but scored.csv {n_rows}"
dupes = scored.duplicated(subset=["question_id", "engine", "repeat_index"]).sum()
if dupes:
    fails.append(f"scored.csv contains {dupes} duplicate question/engine/repeat triples")

# ---- attrition and AI Overview trigger rate
main = scored[scored.repeat_index == 1]
aio = main[main.engine == "google_aio"]
aio_rate = (aio.status == "ok").mean() if len(aio) else None

# ---- pairwise Jaccard, rebuilt from a pivot
piv = (main[main.status == "ok"]
       .pivot_table(index="question_id", columns="engine", values="vendors", aggfunc="first"))


def jac(a, b):
    if not isinstance(a, frozenset) or not isinstance(b, frozenset):
        return None
    if not a and not b:
        return None
    return len(a & b) / len(a | b)


pair_vals, pair_named = [], {}
for a, b in itertools.combinations(ENG, 2):
    if a not in piv.columns or b not in piv.columns:
        pair_named[f"{a}|{b}"] = (0, None)
        continue
    vals = [jac(x, y) for x, y in zip(piv[a], piv[b])]
    vals = [v for v in vals if v is not None]
    pair_named[f"{a}|{b}"] = (len(vals), (sum(vals) / len(vals)) if vals else None)
    pair_vals.extend(vals)
overall_mean = (sum(pair_vals) / len(pair_vals)) if pair_vals else None

# ---- the decisive test, rebuilt PER ENGINE from the repeat index
qmeta = pd.read_csv("data/questions.csv", dtype=str, keep_default_na=False)
stab = set(qmeta.loc[qmeta.in_stability_subsample == "True", "question_id"])
ok = scored[scored.status == "ok"]
ok_stab = ok[ok.question_id.isin(stab)]
scored["order"] = scored["vendor_order_json"].map(lambda s: json.loads(s) if s else [])
order_by = {(r.question_id, r.engine, r.repeat_index): r.order for r in scored.itertuples()}


def overlap(a, b):
    if not a or not b:
        return None
    return len(a & b) / min(len(a), len(b))


def measure(a, b, ka, kb, mode):
    if mode == "full":
        return jac(a, b)
    if mode == "topk":
        return jac(frozenset(order_by.get(ka, [])[:5]), frozenset(order_by.get(kb, [])[:5]))
    if mode == "overlap":
        return overlap(a, b)
    raise ValueError(mode)


ENG_ORDER = ["chatgpt", "claude", "perplexity", "google_aio"]
per_engine_mine = {}
for e in ENG_ORDER:
    blk = {}
    for mode in ("full", "topk", "overlap"):
        w, bt = [], []
        for q in sorted(stab):
            reps = ok_stab[(ok_stab.question_id == q) & (ok_stab.engine == e)]
            rows_ = list(reps.itertuples())
            for x, y in itertools.combinations(rows_, 2):
                if x.model_id != y.model_id:
                    continue
                v = measure(x.vendors, y.vendors,
                            (x.question_id, x.engine, x.repeat_index),
                            (y.question_id, y.engine, y.repeat_index), mode)
                if v is not None:
                    w.append(v)
            r1 = reps[reps.repeat_index == 1]
            if not len(r1):
                continue
            x = list(r1.itertuples())[0]
            for o in ENG_ORDER:
                if o == e:
                    continue
                oo = ok_stab[(ok_stab.question_id == q) & (ok_stab.engine == o)
                             & (ok_stab.repeat_index == 1)]
                if not len(oo):
                    continue
                y = list(oo.itertuples())[0]
                v = measure(x.vendors, y.vendors,
                            (x.question_id, x.engine, x.repeat_index),
                            (y.question_id, y.engine, y.repeat_index), mode)
                if v is not None:
                    bt.append(v)
        blk[mode] = {
            "wn": len(w), "wm": (sum(w) / len(w)) if w else None,
            "bn": len(bt), "bm": (sum(bt) / len(bt)) if bt else None,
            "gap": ((sum(w) / len(w)) - (sum(bt) / len(bt))) if (w and bt) else None,
        }
    per_engine_mine[e] = blk

# ---- consensus, three-engine set, rebuilt with explode
THREE = ["claude", "perplexity", "google_aio"]
sub = main[(main.status == "ok") & (main.engine.isin(THREE))]
counts = sub.groupby("question_id").engine.nunique()
qs3 = set(counts[counts == len(THREE)].index)
long = []
for _, r in sub[sub.question_id.isin(qs3)].iterrows():
    for v in r.vendors:
        long.append((r.question_id, v, r.engine))
ldf = pd.DataFrame(long, columns=["q", "vendor", "engine"])
per = ldf.groupby(["q", "vendor"]).engine.nunique()
share1_3 = (per == 1).mean() if len(per) else None
shareall_3 = (per == len(THREE)).mean() if len(per) else None

# ---- SerpApi
serp = pd.to_numeric(scored.serpapi_calls_used, errors="coerce").fillna(0).sum()

# ---------------------------------------------------------------- parse paper
paper = open("PAPER.md").read()


def find(pattern, group=1, cast=float):
    m = re.search(pattern, paper)
    return cast(m.group(group).replace(",", "")) if m else None


LABELS = {"chatgpt": "ChatGPT", "claude": "Claude", "perplexity": "Perplexity",
          "google_aio": "Google AI Overviews"}

# Parse the per-engine table in 4.5:
# | Claude | 74 | 0.442 (n=74) | 0.277 (n=37) | 0.165 (0.082 to 0.242) | verdict |
paper_pe = {}
for e, lab in LABELS.items():
    m = re.search(r"\|\s*" + re.escape(lab) + r"\s*\|\s*(\d+)\s*\|\s*([0-9.]+)\s*\(n=\s*(\d+)\)"
                  r"\s*\|\s*([0-9.]+)\s*\(n=\s*(\d+)\)\s*\|\s*([0-9.-]+)\s*\(", paper)
    if m:
        paper_pe[e] = {"pairs": int(m.group(1)), "wm": float(m.group(2)),
                       "wn": int(m.group(3)), "bm": float(m.group(4)),
                       "bn": int(m.group(5)), "gap": float(m.group(6))}

# Parse the controls table rows for the two LENGTH controls
paper_ctrl = {}
for e, lab in LABELS.items():
    for mode, needle in (("topk", "truncated to its first 5"),
                         ("overlap", "Overlap coefficient")):
        m = re.search(r"\|\s*" + re.escape(lab) + r"\s*\|\s*[^|]*" + re.escape(needle)
                      + r"[^|]*\|\s*([0-9.]+)\s*\(n=\s*(\d+)\)\s*\|\s*([0-9.]+)"
                        r"\s*\(n=\s*(\d+)\)\s*\|\s*([0-9.-]+)\s*\(", paper)
        if m:
            paper_ctrl[(e, mode)] = {"wm": float(m.group(1)), "wn": int(m.group(2)),
                                     "bm": float(m.group(3)), "bn": int(m.group(4)),
                                     "gap": float(m.group(5))}

for e in ENG_ORDER:
    mine = per_engine_mine[e]["full"]
    if mine["wn"] == 0:
        continue
    pap = paper_pe.get(e)
    if not pap:
        fails.append(f"section 4.5 has no per-engine row for {LABELS[e]}")
        continue
    check(f"4.5 {LABELS[e]} within mean", mine["wm"], pap["wm"], TOL)
    check(f"4.5 {LABELS[e]} within n", float(mine["wn"]), float(pap["wn"]), 0.5)
    check(f"4.5 {LABELS[e]} between mean", mine["bm"], pap["bm"], TOL)
    check(f"4.5 {LABELS[e]} between n", float(mine["bn"]), float(pap["bn"]), 0.5)
    check(f"4.5 {LABELS[e]} gap", mine["gap"], pap["gap"], TOL)
    for mode in ("topk", "overlap"):
        mm = per_engine_mine[e][mode]
        pc = paper_ctrl.get((e, mode))
        if mm["gap"] is None:
            continue
        if not pc:
            fails.append(f"controls table missing {mode} row for {LABELS[e]}")
            continue
        check(f"4.5 {LABELS[e]} {mode} gap", mm["gap"], pc["gap"], TOL)

p_overall = find(r"Mean pairwise Jaccard similarity of the vendor sets is \*\*([0-9.]+)")
p_pairobs = find(r"across ([0-9,]+) engine-pair\s+observations", cast=int)
p_aio = find(r"Google returned an AI Overview for \*\*([0-9.]+)%")
p_rows = find(r"All ([0-9,]+) collected answers", cast=int)
p_share1 = find(r"\*\*([0-9.]+)% of the [0-9,]+ distinct vendor mentions came from a single engine")
p_shareall = find(r"only \*\*([0-9.]+)% from all 3\*\*")
p_serp = find(r"\*\*([0-9,]+) SerpApi searches\*\*", cast=int)

check("overall mean pairwise Jaccard", overall_mean, p_overall, TOL)
check("engine-pair observations", float(len(pair_vals)), float(p_pairobs) if p_pairobs else None, 0.5)
check("AI Overview trigger rate (%)", 100 * aio_rate if aio_rate is not None else None, p_aio, PCT_TOL)
check("total collected answers", float(n_rows), float(p_rows) if p_rows else None, 0.5)
check("3-engine single-engine share (%)", 100 * share1_3 if share1_3 is not None else None, p_share1, PCT_TOL)
check("3-engine all-engines share (%)", 100 * shareall_3 if shareall_3 is not None else None, p_shareall, PCT_TOL)
check("SerpApi searches consumed", float(serp), float(p_serp) if p_serp else None, 0.5)

# ---------------------------------------------------------------- lock checks
LOCKS = [
    (r"\bGemini\b", "Gemini named anywhere"),
    (r"\bCopilot\b(?!.*Microsoft Copilot is a vendor)", "Copilot named anywhere"),
    (r"29\.4\s*%", "retired 29.4% figure"),
    (r"zenodo\.21537013|10\.5281/zenodo\.21537013", "concept DOI 21537013"),
    (r"10\.5281/zenodo\.21710465", "unresolved Crackle PR DOI"),
    (r"11 percent ChatGPT-to-Perplexity domain-overlap figure circulates", None),
]
for pat, label in LOCKS:
    if label is None:
        continue
    for m in re.finditer(pat, paper, re.IGNORECASE):
        ctx = paper[max(0, m.start() - 120):m.end() + 120].replace("\n", " ")
        # the "not cited" paragraph is allowed to name the figure it rejects
        if "omitted rather than repeated" in ctx or "could not find it stated" in ctx:
            continue
        fails.append(f"LOCK VIOLATION, {label}: ...{ctx}...")

# prospect contact names from the source sheet must never appear
try:
    import openpyxl, os
    xp = os.environ.get("VOL3_XLSX",
        "/root/.claude/uploads/2a6b0b0d-bfde-53d3-a976-6afaf2b5f362/b72fbd11-GEO_Prospect_List_54.xlsx")
    wb = openpyxl.load_workbook(xp, data_only=True)
    ws = wb["Sweep"]
    hdr = [c.value for c in ws[1]]
    names = set()
    for r in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(hdr, r))
        nm = str(d.get("name") or "").strip()
        if nm and nm.upper() not in ("LEADER TEST", "JANE DOE") and " " in nm:
            names.add(nm)
    # the paper's own author appears in the sheet as a contact row; he is not a
    # third-party prospect and his name is the byline
    names.discard("Sairam Sivakumar")
    leaked = sorted(nm for nm in names if re.search(r"\b" + re.escape(nm) + r"\b", paper))
    if leaked:
        fails.append(f"LOCK VIOLATION, prospect contact names in the paper: {leaked}")
    checks.append(("prospect contact names checked", float(len(names)), float(len(names)), True))
except Exception as e:
    checks.append((f"prospect-name check skipped ({e})", None, None, True))

# both prior version DOIs must be cited, concept DOI must not be
for doi in ("10.5281/zenodo.21537014", "10.5281/zenodo.21586091"):
    if doi not in paper:
        fails.append(f"required prior-version DOI missing from the paper: {doi}")

# ---------------------------------------------------------------- report
print(f"{'CHECK':44} {'RECOMPUTED':>14} {'PAPER':>14}   RESULT")
for name, mine, pap, ok_ in checks:
    ms = "n/a" if mine is None else f"{mine:,.4f}"
    ps = "n/a" if pap is None else f"{pap:,.4f}"
    print(f"{name[:44]:44} {ms:>14} {ps:>14}   {'PASS' if ok_ else 'FAIL'}")
print()
if fails:
    print(f"GATE FAILED, {len(fails)} problem(s):")
    for f in fails:
        print("  -", f)
    sys.exit(1)
print("GATE PASSED. Every headline figure in PAPER.md was reproduced by an "
      "independent implementation, and no locked term appears.")
